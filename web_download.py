# -*- coding: utf-8 -*-
"""
需求管理平台 (RMP) 语料下载：
  python web_download.py login              # 弹浏览器手动登录一次，保存会话
  python web_download.py download           # 一键下载列表(待处理)全部语料到 _inbox
  python web_download.py download --base <RMP_BASE>
规则：每个需求优先下【资源汇总】，没有则【逆规整后】，按单号(seqNo)命名。
内网地址配置在同目录 endpoints.json（参见 endpoints.example.json）。
"""
import os
import re
import sys
import argparse
from urllib.parse import urlsplit


def app_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


try:  # GBK 控制台下 ✓/↓ 等字符会崩，统一 UTF-8 输出
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

APP = app_dir()
SESSION_DIR = os.path.join(APP, "_session")
STATE = os.path.join(SESSION_DIR, "storage_state.json")
INBOX = os.path.join(APP, "_inbox")


def _endpoint(key, default=""):
    """内网地址：优先环境变量，其次同目录 endpoints.json；公开仓库不含真实地址。"""
    import json as _json
    v = os.environ.get(key, "").strip()
    if v:
        return v
    try:
        _base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
            else os.path.dirname(os.path.abspath(__file__))
    except Exception:
        _base = os.getcwd()
    for _b in (_base, os.getcwd()):
        try:
            _p = os.path.join(_b, "endpoints.json")
            if os.path.exists(_p):
                _d = _json.load(open(_p, encoding="utf-8"))
                if _d.get(key):
                    return str(_d[key])
        except Exception:
            pass
    return default


RMP_BASE = _endpoint("RMP_BASE")
LIST_PATH = "/offer-page/#/requirement"
SEARCH_API = "/rmp/v2/bug/requirement/search"
ATTACH_API = "/rmp/v2/bug/requirement/comment/getAttachList"
BASE_BODY = {"keyWord": "", "purityStatus": "", "operatorStatus": "0",
             "requirementType": "", "status": "", "startTime": "", "endTime": "",
             "cloudNativeFlag": "", "language": "", "createUser": ""}
PICK_PRIORITY = ["资源汇总", "逆规整后"]
PAGE_SIZE = 50


def _safe(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", str(name)).strip()


SCHEDULE = os.path.join(APP, "排期.xlsx")
# 语种代码 -> 排期用中文目录名（含大小写变体）
LANG_CN = {
    'ar_il': '阿语', 'de': '德语', 'fr_fr': '法语', 'he_il': '希伯来语',
    'pt_la': '葡语', 'pt_pt': '葡萄牙语', 'th_th': '泰语', 'da_dk': '丹麦语',
    'nb_no': '挪威语', 'sv_se': '瑞典语', 'nl_nl': '荷兰语', 'hu': '匈牙利语',
    'ru': '俄语', 'si_si': '斯洛文尼亚语', 'vi_vn': '越南语', 'ko_kr': '韩语',
    'en': '英语', 'en_uk': '英语', 'en_au': '英语', 'en_ml': '英语',
    'ja': '日语', 'es_la': '西班牙语', 'es_es': '西班牙语', 'it_it': '意大利语',
    'pl_pl': '波兰语', 'tr_tr': '土耳其语', 'fa_ir': '波斯语', 'hi_in': '印地语',
    'id_id': '印尼语', 'ms_my': '马来语',
}


def _lang_cn(seq_no: str) -> str:
    """从单号后缀（...-ASR-th_th）反查中文语种名；非标单号无后缀则空。"""
    m = re.search(r'ASR-([A-Za-z_]+)$', seq_no or '')
    if not m:
        return ''
    return LANG_CN.get(m.group(1).lower(), '')


def _lang_aliases(v) -> set:
    """语种别名集合（西语↔西班牙语↔es_la）。表在 pipeline 里，这边只借用，不复制——
    复制一份迟早两边改不同步，那正是我们一路在修的『副本分叉』。"""
    try:
        from pipeline import _lang_aliases as _pa
        return _pa(v)
    except Exception:
        s = str(v or '').strip().lower().replace(' ', '')
        return {s} if s else set()


def _brand_of(item: dict) -> str:
    """从搜索项里取车厂：companyName 优先，其次项目名前段。"""
    v = item.get('companyName')
    if v:
        return str(v).strip()
    arr = item.get('reqProjects') or item.get('projects')
    if isinstance(arr, list) and arr:
        v = arr[0].get('companyName') or arr[0].get('projectName')
        if v:
            return str(v).split('-')[0].strip()
    pn = item.get('projectName')
    if pn:
        return str(pn).split('-')[0].strip()
    return ''


def update_schedule(rows: list) -> int:
    """把 (单号,车厂,语种,预计完成) 追加进排期表（去重，不覆盖已有）。返回新增行数。"""
    if not rows:
        return 0
    try:
        from openpyxl import load_workbook, Workbook
        if os.path.exists(SCHEDULE):
            wb = load_workbook(SCHEDULE)
            ws = wb.active
            existing = {str(r[0]).strip() for r in
                        ws.iter_rows(min_row=2, max_col=1, values_only=True) if r[0]}
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["母任务", "车厂", "语种", "预计完成时间"])
            existing = set()
        added = 0
        for seq, brand, lang, deadline in rows:
            if seq and seq not in existing:
                ws.append([seq, brand, lang, deadline])
                existing.add(seq)
                added += 1
        wb.save(SCHEDULE)
        return added
    except Exception as e:
        # 返回 0 会被上层打成"✓ 新增 0 行（已存在的不重复）"——写失败和本来就都有，
        # 长得一模一样。用 None 区分开。最常见的失败：排期.xlsx 正被 Excel 打开。
        print(f"  [排期] ✗ 写入失败：{e}", flush=True)
        return None


def upsert_schedule(rows: list) -> tuple:
    """把 (单号,车厂,语种,预计完成) 更新进排期表：已存在则用文档的【非空】车厂/语种覆盖，
    不存在则追加（用于「刷新排期」——以腾讯文档为准）。返回 (更新数, 新增数)。"""
    if not rows:
        return (0, 0)
    try:
        from openpyxl import load_workbook, Workbook
        if os.path.exists(SCHEDULE):
            wb = load_workbook(SCHEDULE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["母任务", "车厂", "语种", "预计完成时间"])
        idx = {}
        for i, r in enumerate(ws.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if r[0]:
                idx[str(r[0]).strip()] = i
        updated = added = 0
        for seq, brand, lang, deadline in rows:
            seq = str(seq or "").strip()
            if not seq:
                continue
            if seq in idx:
                ri = idx[seq]
                changed = False
                if str(brand or "").strip() and \
                        str(ws.cell(ri, 2).value or "").strip() != str(brand).strip():
                    ws.cell(ri, 2).value = brand
                    changed = True
                if str(lang or "").strip() and \
                        str(ws.cell(ri, 3).value or "").strip() != str(lang).strip():
                    ws.cell(ri, 3).value = lang
                    changed = True
                if changed:
                    updated += 1
            else:
                ws.append([seq, brand, lang, deadline])
                idx[seq] = ws.max_row
                added += 1
        wb.save(SCHEDULE)
        return (updated, added)
    except Exception as e:
        print(f"  [排期] 刷新失败：{e}", flush=True)
        return (0, 0)


def login(base: str = RMP_BASE) -> int:
    """弹出浏览器，用户手动登录，会话持久化到 storage_state.json。完成后关闭窗口即可。"""
    from playwright.sync_api import sync_playwright
    os.makedirs(SESSION_DIR, exist_ok=True)
    url = base.rstrip("/") + LIST_PATH
    print("正在打开浏览器…请登录后【关闭浏览器窗口】完成。", flush=True)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, args=["--start-maximized"])
        ctx = browser.new_context(no_viewport=True, accept_downloads=True)
        page = ctx.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60000)
        except Exception as e:
            print(f"[提示] 打开页面异常（不影响登录）：{e}", flush=True)
        i = 0
        while True:
            try:
                page.wait_for_timeout(1000)
                if not ctx.pages:
                    break
                i += 1
                if i % 5 == 0:
                    ctx.storage_state(path=STATE)
            except Exception:
                break
        try:
            ctx.storage_state(path=STATE)
        except Exception:
            pass
        try:
            browser.close()
        except Exception:
            pass
    ok = os.path.exists(STATE)
    print("✓ 登录会话已保存。" if ok else "✗ 未捕获到会话，请重试。", flush=True)
    return 0 if ok else 1


def _pick(attaches):
    for kw in PICK_PRIORITY:
        for a in attaches:
            if kw in str(a.get("name", "")) and str(a.get("size") or "0") != "0":
                return a
    return None


def download(base: str = RMP_BASE, scope: str = "all",
             lang: str = "", brand: str = "") -> int:
    """
    遍历列表(待处理)需求，下载语料到 _inbox，按单号命名。
    scope: all=全部 / local=只本地(-L-ASR) / cloud=只云端(-C-ASR)
    lang/brand: 可选，只下载匹配语种/车厂的单（留空=不限）。
    """
    lang = (lang or os.environ.get("FILTER_LANG", "")).strip()
    brand = (brand or os.environ.get("FILTER_BRAND", "")).strip()
    from playwright.sync_api import sync_playwright
    if not os.path.exists(STATE):
        print("✗ 未找到登录会话，请先点“登录”。", flush=True)
        return 2
    os.makedirs(INBOX, exist_ok=True)
    base = base.rstrip("/")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(storage_state=STATE)
        req = ctx.request

        items, page, total = [], 1, None
        while True:
            r = req.post(base + SEARCH_API, data=dict(BASE_BODY, pageNo=page, pageSize=PAGE_SIZE),
                         timeout=40000)
            if r.status != 200:
                print(f"✗ 列表请求 HTTP {r.status}（会话可能失效，请重新登录）", flush=True)
                browser.close()
                return 3
            data = (r.json() or {}).get("data", {}) or {}
            total = data.get("totalSize", total)
            content = data.get("content", []) or []
            items.extend(content)
            if not content or (total is not None and len(items) >= total) or page > 80:
                break
            page += 1
        if not items:
            print("✗ 列表为空（会话可能失效，请重新登录）", flush=True)
            browser.close()
            return 3

        # 云端/本地同名去重：本地有同名单时跳过云端（云端独有的保留）
        # 规则1 同号仅 -L-/-C- 不同；规则2 NS 单编号差±1 且 标题(去单号)一致 或 语种+车厂一致
        def _norm(s):
            return re.sub(r'-(?:L|C)-ASR', '-ASR', str(s or ''))

        def _num(s):
            m = re.search(r'(\d{4,})', str(s or ''))
            return int(m.group(1)) if m else None

        def _tkey(it):
            t = str(it.get("title") or '')
            t = re.sub(r'(?:N-CUS|NS)-\d+(?:-[A-Za-z_]+)*', '', t)
            return re.sub(r'[\s　]+', '', t)

        locals_norm = set()
        locals_seqs = set()
        locals_ns = []
        for it in items:
            seq = str(it.get("seqNo", ""))
            if "-L-ASR" in seq:
                locals_norm.add(_norm(seq))
                locals_seqs.add(seq)
                if seq.startswith("NS-"):
                    locals_ns.append((_num(seq), _tkey(it),
                                      str(it.get("languageName") or ''), _brand_of(it)))

        def _is_dup(it):
            seq = str(it.get("seqNo", ""))
            if "-C-ASR" not in seq:
                return False
            if _norm(seq) in locals_norm:
                return True
            # 规则3：标题里直接写着某个本地单号
            refs = re.findall(r'(?:N-CUS|NS)-\d+(?:-[A-Za-z_]+)*', str(it.get("title") or ''))
            if any(x in locals_seqs for x in refs):
                return True
            if seq.startswith("NS-"):
                num, tk = _num(seq), _tkey(it)
                lang, brand = str(it.get("languageName") or ''), _brand_of(it)
                for n2, tk2, l2, b2 in locals_ns:
                    if num is None or n2 is None or abs(num - n2) != 1:
                        continue
                    if (tk and tk == tk2) or (lang and brand and lang == l2 and brand == b2):
                        return True
            return False

        dup_skip = [str(it.get("seqNo")) for it in items if _is_dup(it)]
        if dup_skip:
            items = [it for it in items if not _is_dup(it)]
            print(f"云端/本地同名去重：跳过 {len(dup_skip)} 个云端单（本地已有同名/相邻同名）", flush=True)
            for s in dup_skip:
                print(f"  · {s}", flush=True)

        if scope == "local":
            items = [it for it in items if "-L-ASR" in str(it.get("seqNo", ""))]
        elif scope == "cloud":
            items = [it for it in items if "-C-ASR" in str(it.get("seqNo", ""))]
        # 语种/车厂筛选（留空=不限；模糊：大小写不敏感、含即命中）
        def _fz(needle, *fields):
            n = str(needle).lower().replace(" ", "")
            return any(n in str(f or "").lower().replace(" ", "") for f in fields)
        if lang:
            # 语种可用 中文名 / 单号(-ASR-代码) / languageName 任意片段匹配。
            # 走别名集合：填"西语"要能命中"西班牙语"（纯子串匹配会漏）。
            _aliases = _lang_aliases(lang)
            items = [it for it in items
                     if any(_fz(a, _lang_cn(str(it.get("seqNo", ""))),
                                it.get("languageName"), it.get("seqNo"))
                            for a in _aliases)]
        if brand:
            items = [it for it in items if _fz(brand, _brand_of(it))]
        scope_cn = {"local": "只本地", "cloud": "只云端"}.get(scope, "全部")
        flt_cn = (f"，语种[{lang}]" if lang else "") + (f"，车厂[{brand}]" if brand else "")
        print(f"列表共 {total} 条，范围[{scope_cn}]{flt_cn} 命中 {len(items)} 条，开始下载…", flush=True)
        ok = skip = fail = 0
        sched_rows = []
        for i, it in enumerate(items, 1):
            rid = str(it.get("id")); seq = it.get("seqNo")
            try:
                r = req.get(f"{base}{ATTACH_API}?reqId={rid}&pageNo=1&pageSize=9999", timeout=30000)
                atts = (r.json() or {}).get("data", {}).get("content", []) or []
            except Exception as e:
                print(f"[{i}/{len(items)}] {seq} 附件列表失败：{e}", flush=True)
                fail += 1
                continue
            picked = _pick(atts)
            if not picked:
                print(f"[{i}/{len(items)}] {seq} 无资源汇总/逆规整后，跳过", flush=True)
                skip += 1
                continue
            fname = f"{_safe(seq)}.xlsx"
            try:
                r = req.get(picked["path"], timeout=120000)
                body = r.body()
                # 会话中途过期时接口会返回 200 的登录页/错误 JSON，.body() 照样给字节。
                # 不校验就会把 HTML 当 xlsx 落盘，还一路被归档进语料库。
                bad = ''
                if r.status != 200:
                    bad = f"HTTP {r.status}"
                elif not body[:2] == b'PK':          # xlsx 是 zip，magic 必是 PK
                    head = body[:200].decode('utf-8', 'replace').strip().replace('\n', ' ')
                    bad = f"返回的不是 xlsx（{len(body)}B，开头：{head[:80]}）"
                if bad:
                    print(f"[{i}/{len(items)}] {seq} 下载失败：{bad}", flush=True)
                    fail += 1
                    continue
                with open(os.path.join(INBOX, fname), "wb") as f:
                    f.write(body)
                print(f"[{i}/{len(items)}] {seq}  ↓ {len(body)//1024}KB  {fname}", flush=True)
                ok += 1
                brand = _brand_of(it)
                if not brand:   # 搜索项无车厂时，回退到详情接口
                    try:
                        d = (req.get(f"{base}/rmp/v2/cust/requirement/getCustRequirementInfo?rId={rid}",
                                     timeout=20000).json() or {}).get("data", {}) or {}
                        brand = (d.get("companyName") or "").strip()
                        if not brand:
                            projs = d.get("projects") or []
                            if projs:
                                brand = (projs[0].get("companyName") or "").strip()
                    except Exception:
                        pass
                sched_rows.append((seq, brand, _lang_cn(seq), it.get("deadline") or ""))
            except Exception as e:
                print(f"[{i}/{len(items)}] {seq} 下载失败：{e}", flush=True)
                fail += 1
        browser.close()
        added = update_schedule(sched_rows)
        mark = '✓' if not fail else '✗'
        print(f"{mark} 完成：下载 {ok}，跳过 {skip}，失败 {fail}  ->  _inbox", flush=True)
        if added is None:
            print("✗ 排期表：写入失败，本批单号没进排期（排期.xlsx 是不是正被 Excel 打开？）"
                  "  —— 不补录的话，处理时这些文件会全进 待确认/", flush=True)
        else:
            print(f"✓ 排期表：自动新增 {added} 行（已存在的不重复）  ->  排期.xlsx", flush=True)
    return 1 if (fail or added is None) else 0


def _origin(url: str) -> str:
    """把可能带 /offer-page/#/... 路径的整页地址归一化为源站 scheme://host。
    这样用户在地址栏粘贴哪个页面 URL 都能用，避免拼接 API 时拼出畸形地址(→HTTP405)。"""
    if not url:
        return url
    p = urlsplit(url.strip())
    if p.scheme and p.netloc:
        return f"{p.scheme}://{p.netloc}"
    return url.strip().rstrip("/")


def main():
    ap = argparse.ArgumentParser(description="RMP 语料下载")
    ap.add_argument("cmd", choices=["login", "download"])
    ap.add_argument("--base", default=RMP_BASE)
    ap.add_argument("--scope", default="all", choices=["all", "local", "cloud"])
    ap.add_argument("--lang", default="")
    ap.add_argument("--brand", default="")
    args = ap.parse_args()
    base = _origin(args.base)                 # 归一化到源站，容忍粘贴整页地址
    if args.cmd == "login":
        sys.exit(login(base))
    else:
        sys.exit(download(base, args.scope, args.lang, args.brand))


if __name__ == "__main__":
    main()
