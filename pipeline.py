"""
ASR 语料整理流水线
==================
用法:
  python pipeline.py              # 正常运行（从 _inbox/ 或根目录取文件）
  python pipeline.py --send       # 正常运行并在结尾把本批 new_corpus+测试集 传到 rdg
  python pipeline.py restore      # 从最新备份恢复并重新处理
  python pipeline.py restore <目录>  # 从指定备份恢复
  python pipeline.py clean        # 全量标准化历史文件
  python pipeline.py package      # 只重新打包 ZIP
  python pipeline.py verify       # 只跑核验，不移动文件
  python pipeline.py send         # 只重发最近一批产物（rftctl -> rdg）
  python pipeline.py send --file <路径>   # 传指定文件/文件夹
可选参数:
  --receive <网段[,网段...]>          接收网段，逗号分隔可多发（默认 rdg,dtn；或环境变量 RFT_RECEIVE）
  --keep N                            各类产物保留份数（默认 5）
  --no-tts                            跳过 TTS 测试集合成（Step 4）
"""
import os, re, sys, shutil, zipfile, subprocess, random, glob, time
from datetime import datetime

try:  # GBK 控制台下 ✓/→ 等字符会崩，统一 UTF-8 输出
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
from openpyxl import load_workbook
import pandas as pd

# ── 常量 ──────────────────────────────────────────────
SCHEDULE_FILE   = '排期.xlsx'
INBOX_DIR       = '_inbox'           # 新文件放这里
INCREMENTAL_DIR_PREFIX = '新增语料'
BACKUP_DIR_PREFIX = 'backup_语料'
BACKUPS_DIR     = '_backups'          # 备份归拢目录
OUTPUTS_DIR     = '_outputs'          # 增量语料 / zip 输出
REPORTS_DIR     = '_reports'          # 处理报告 / 排期状态
ARCHIVE_DIR     = '_archive'          # 历史产物归档（人工清理）
KEEP_RECENT     = int(os.environ.get('PIPELINE_KEEP', '5'))  # 各类产物保留份数
INTERNAL_DIRS   = {INBOX_DIR, BACKUPS_DIR, OUTPUTS_DIR, REPORTS_DIR,
                   ARCHIVE_DIR, '__pycache__', '待确认'}
# 打包/构建产物目录：不是语料，扫描时一律跳过
BUILD_ARTIFACT_DIRS     = {'dist', 'build'}
BUILD_ARTIFACT_PREFIXES = ('asr_pipeline_',)
def _resolve_tts_tools_dir():
    """定位 TTS 工具目录：环境变量 > 程序目录/tts_tools（打包/OTA 后在这）> D:\\tts_tools（开发机兜底）。"""
    v = os.environ.get('TTS_TOOLS_DIR', '').strip()
    if v:
        return v
    try:
        base = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) \
            else os.path.dirname(os.path.abspath(__file__))
    except Exception:
        base = os.getcwd()
    for cand in (os.path.join(base, 'tts_tools'),
                 os.path.join(os.getcwd(), 'tts_tools'),
                 r'D:\tts_tools'):
        if os.path.isfile(os.path.join(cand, 'asr_tts_tool.py')):
            return cand
    return r'D:\tts_tools'


TTS_TOOLS_DIR   = _resolve_tts_tools_dir()
TTS_SAMPLE_SIZE = 1000
TTS_CONCURRENCY = 10
# ── rftctl 跨网段传输 ──────────────────────────────
def _endpoint(key, default=""):
    """内网地址/路径：优先环境变量，其次同目录 endpoints.json；公开仓库不含真实值。"""
    import json as _json
    v = os.environ.get(key, "").strip()
    if v:
        return v
    try:
        _base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) \
            else os.path.dirname(os.path.abspath(__file__))
    except Exception:
        _base = os.getcwd()
    for _b in (_base, os.getcwd()):
        try:
            _p = os.path.join(_b, "endpoints.json")
            if os.path.exists(_p):
                _d = _json.load(open(_p, encoding="utf-8"))
                if _d.get(key):
                    return str(_d[key])
        except Exception:
            pass
    return default


RFTCTL_PATH_DEFAULT = _endpoint("RFTCTL_PATH", "")
# 接收网段，逗号分隔可多发（默认同时发 rdg 和 dtn）
TRANSFER_RECEIVES   = os.environ.get('RFT_RECEIVE', 'rdg,dtn')
TODAY_TIMESTAMP = datetime.now().strftime('%Y%m%d')
RUN_TIMESTAMP   = datetime.now().strftime('%Y%m%d_%H%M%S')
TIMESTAMP_REGEX = r'_\d{8}'
CHINESE_PATTERN = re.compile(r'[\u4e00-\u9fa5]')
DEFAULT_BRANDS  = ['东风', '比亚迪', '吉利', '极氪', '奇瑞', '长安', '长城']
BRAND_ALIASES   = {'领跑': '零跑'}
ZIP_BRAND_NAMES = {
    '\u4e1c\u98ce': 'dongfeng',
    '\u516c\u7248': 'gongban',
    '\u5317\u6c7d': 'beiqi',
    '\u5409\u5229': 'jili',
    '\u5947\u745e': 'qirui',
    '\u6781\u6c2a': 'zeekr',
    '\u6bd4\u4e9a\u8fea': 'byd',
    '\u7ea2\u65d7': 'hongqi',
    '\u957f\u57ce': 'changcheng',
    '\u957f\u5b89': 'changan',
    '\u963f\u7ef4\u5854': 'avatr',
    '\u96f6\u8dd1': 'leapmotor',
    '\u9886\u8dd1': 'leapmotor',
}
ZIP_LANG_NAMES = {
    '\u963f\u8bed': 'ar_il',
    '\u5fb7\u8bed': 'de',
    '\u6cd5\u8bed': 'fr_fr',
    '\u5e0c\u4f2f\u6765\u8bed': 'he_il',
    '\u8461\u8bed': 'pt_la',
    '\u8461\u8404\u7259\u8bed': 'pt_la',
    '\u6cf0\u8bed': 'th_th',
    '\u4e39\u9ea6\u8bed': 'da_dk',
    '\u632a\u5a01\u8bed': 'nb_no',
    '\u745e\u5178\u8bed': 'sv_se',
    '\u8377\u5170\u8bed': 'nl_nl',
    '\u5308\u7259\u5229\u8bed': 'hu',
    '\u4fc4\u8bed': 'ru',
}

SHUOFA_HARD_KEYWORDS     = ["扣槽", "抠槽", "鎶犳Ы", "_shuofa", "槽合集"]
SHUOFA_FALLBACK_KEYWORDS = ["功能需求"]   # 仅在无 sent sheet 时才视为 shuofa
SHUOFA_KEYWORDS          = SHUOFA_HARD_KEYWORDS + SHUOFA_FALLBACK_KEYWORDS
SENT_KEYWORDS            = ["语料", "逆规整", "句子", "鍏ㄩ噺", "閫嗚", "_sent"]

# ── 内容特征词（不依赖 sheet 名，直接看 header 行）────────
# 出现 2+ 个 → 该 sheet 是管理/元数据表，直接丢弃
_META_SIGNALS = ['一级功能', '二级功能', '三级功能', '语义取值',
                 '文档密级', '是否纳管', '纳管类型', '生效条件']
# header 行含以下关键词 → 该列是逆规整（sent）目标列
_SENT_COL_KW  = ['逆规整后', '逆规整列', '逆规整', '閫嗚']

_LATIN = r'[a-zA-Z]'
LANG_VALIDATORS = {
    'ar_il': lambda s: any('\u0600' <= c <= '\u06ff' for c in s),
    'de':    lambda s: bool(re.search(_LATIN + r'|[äöüÄÖÜß]', s)),
    'fr_fr': lambda s: bool(re.search(_LATIN + r'|[àâéèêëîïôùûüçÀÂÉÈÊËÎÏÔÙÛÜÇ]', s)),
    'th_th': lambda s: any('\u0e00' <= c <= '\u0e7f' for c in s),
    'da_dk': lambda s: bool(re.search(_LATIN + r'|[æøåÆØÅ]', s)),
    'nb_no': lambda s: bool(re.search(_LATIN + r'|[æøåÆØÅ]', s)),
    'sv_se': lambda s: bool(re.search(_LATIN + r'|[åäöÅÄÖ]', s)),
    'nl_nl': lambda s: bool(re.search(_LATIN, s)),
    'si_si': lambda s: bool(re.search(_LATIN + r'|[čšžČŠŽ]', s)),
    'en':    lambda s: bool(re.search(_LATIN, s)),
    'en_uk': lambda s: bool(re.search(_LATIN, s)),
    'en_au': lambda s: bool(re.search(_LATIN, s)),
    'en_ml': lambda s: bool(re.search(_LATIN, s)),
    'ms_my': lambda s: bool(re.search(_LATIN, s)),
    'id_id': lambda s: bool(re.search(_LATIN, s)),
    'vi_vn': lambda s: bool(re.search(_LATIN + r'|[àáảãạăâđêôơư]', s)),
    'es_la': lambda s: bool(re.search(_LATIN + r'|[áéíóúñ¿¡ÁÉÍÓÚÑ]', s)),
    'es_es': lambda s: bool(re.search(_LATIN + r'|[áéíóúñ¿¡ÁÉÍÓÚÑ]', s)),
    'it_it': lambda s: bool(re.search(_LATIN + r'|[àèéìòùÀÈÉÌÒÙ]', s)),
    'pl_pl': lambda s: bool(re.search(_LATIN + r'|[ąćęłńóśźżĄĆĘŁŃÓŚŹŻ]', s)),
    'tr_tr': lambda s: bool(re.search(_LATIN + r'|[çğışöüÇĞİŞÖÜ]', s)),
    'pt_la': lambda s: bool(re.search(_LATIN + r'|[ãõáéíóúâêôàçÁÉÍÓÚ]', s)),
    'pt_pt': lambda s: bool(re.search(_LATIN + r'|[ãõáéíóúâêôàçÁÉÍÓÚ]', s)),
    'fa_ir': lambda s: any('؀' <= c <= 'ۿ' for c in s),
    'hi_in': lambda s: any('ऀ' <= c <= 'ॿ' for c in s),
    'ja':    lambda s: any(('぀' <= c <= 'ヿ') or ('一' <= c <= '鿿') for c in s),
    'ko_kr': lambda s: any('가' <= c <= '힯' for c in s),
    'ru':    lambda s: any('Ѐ' <= c <= 'ӿ' for c in s),
}
DIR_LANG_VALIDATORS = {
    '阿语':          LANG_VALIDATORS['ar_il'],
    '德语':          LANG_VALIDATORS['de'],
    '法语':          LANG_VALIDATORS['fr_fr'],
    '泰语':          LANG_VALIDATORS['th_th'],
    '丹麦语':        LANG_VALIDATORS['da_dk'],
    '挪威语':        LANG_VALIDATORS['nb_no'],
    '瑞典语':        LANG_VALIDATORS['sv_se'],
    '荷兰语':        LANG_VALIDATORS['nl_nl'],
    '斯洛文尼亚语':  LANG_VALIDATORS['si_si'],
}

TTS_LANG_LOCALES = {
    'ar_il': 'ar-SA',
    'de': 'de-DE',
    'fr_fr': 'fr-FR',
    'he_il': 'he-IL',
    'pt_la': 'pt-PT',
    'th_th': 'th-TH',
    'da_dk': 'da-DK',
    'nb_no': 'nb-NO',
    'sv_se': 'sv-SE',
    'nl_nl': 'nl-NL',
    'hu': 'hu-HU',
    'ru': 'ru-RU',
    'si_si': 'sl-SI',
    'en': 'en-US',
    'en_uk': 'en-GB',
    'en_au': 'en-AU',
    'en_ml': 'en-US',
    'ja': 'ja-JP',
    'ko_kr': 'ko-KR',
    'es_la': 'es-MX',
    'es_es': 'es-ES',
    'it_it': 'it-IT',
    'pl_pl': 'pl-PL',
    'tr_tr': 'tr-TR',
    'fa_ir': 'fa-IR',
    'hi_in': 'hi-IN',
    'id_id': 'id-ID',
    'vi_vn': 'vi-VN',
    'ms_my': 'ms-MY',
    'pt_pt': 'pt-PT',
}
DIR_TTS_LANGS = {
    '德语': 'de',
    '法语': 'fr_fr',
    '泰语': 'th_th',
    '丹麦语': 'da_dk',
    '挪威语': 'nb_no',
    '瑞典语': 'sv_se',
    '荷兰语': 'nl_nl',
    '匈牙利语': 'hu',
    '俄语': 'ru',
    '斯洛文尼亚语': 'si_si',
}


# ══════════════════════════════════════════════════════
# 内部工具
# ══════════════════════════════════════════════════════

def _is_shuofa_hard(name):     return any(kw in name for kw in SHUOFA_HARD_KEYWORDS)
def _is_shuofa_fallback(name): return any(kw in name for kw in SHUOFA_FALLBACK_KEYWORDS)
def _is_shuofa(name):          return _is_shuofa_hard(name) or _is_shuofa_fallback(name)
def _is_sent(name):            return any(kw in name for kw in SENT_KEYWORDS)
def _normalize_brand(name):
    s = _cell_str(name)                       # 空/NaN → ''，不再变成 'nan'
    return BRAND_ALIASES.get(s, s)

def _sheet_is_meta(ws) -> bool:
    """
    扫描 sheet 前 3 行的 header，若命中 2+ 个管理表特征词 → 元数据表，应丢弃。
    这样无论 sheet 名是什么（乱码/非标），都能正确识别。
    """
    vals = []
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=25, values_only=True):
        vals.extend(str(v) for v in row if v is not None)
    txt = ' '.join(vals)
    return sum(1 for kw in _META_SIGNALS if kw in txt) >= 2

def _find_sent_col(ws):
    """
    在 header 行找逆规整目标列，返回 1-based 列号；未找到返回 None。
    """
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return None
    for i, v in enumerate(header, 1):
        if v and any(kw in str(v) for kw in _SENT_COL_KW):
            return i
    return None

def _shuofa_content_ok(ws) -> bool:
    """
    shuofa 应当是目标语言的句式，不应有大量中文正文。
    取前 30 个非空数据单元格，若中文字符占比 > 40% → 拒绝。
    """
    samples = []
    for row in ws.iter_rows(min_row=2, max_row=50, max_col=10, values_only=True):
        for v in row:
            s = str(v or '').strip()
            if len(s) > 1:
                samples.append(s)
            if len(samples) >= 30:
                break
        if len(samples) >= 30:
            break
    if not samples:
        return True
    chinese_ratio = sum(
        1 for s in samples
        if sum(1 for c in s if '\u4e00' <= c <= '\u9fa5') / max(len(s), 1) > 0.3
    ) / len(samples)
    return chinese_ratio <= 0.4

def _file_lang(name_no_ext: str) -> str:
    base = re.sub(TIMESTAMP_REGEX, '', name_no_ext)
    m = re.search(r'ASR-(.+)$', base)
    return m.group(1).lower() if m else ''

def _cell_str(v) -> str:
    """把单元格值转成字符串；空/NaN → ''（避免 pandas 的 NaN 变成字面 'nan'）。"""
    try:
        if v is None or (isinstance(v, float) and v != v):   # NaN != NaN
            return ''
    except Exception:
        pass
    s = str(v).strip()
    return '' if s.lower() == 'nan' else s


def _build_mapping():
    """从排期.xlsx 构建完整映射列表。"""
    if not os.path.exists(SCHEDULE_FILE):
        return []
    df = pd.read_excel(SCHEDULE_FILE)
    mapping = []
    for _, row in df.iterrows():
        task_id = str(row.iloc[0]).strip()
        m = re.search(r'\d{4,}', task_id)
        if not m:
            continue
        core_id   = m.group()
        lm        = re.search(r'ASR-(.+)$', task_id)
        task_lang = lm.group(1).lower() if lm else ''
        brand     = _normalize_brand(row.iloc[1])
        lang_dir  = _cell_str(row.iloc[2])          # 空单元格→''，不再变成 'nan'
        mapping.append((core_id, task_lang, brand, lang_dir, task_id))
    return mapping

def _get_brands(mapping: list = None) -> list:
    """从默认配置、排期和当前目录自动汇总车厂。"""
    brands = set(DEFAULT_BRANDS)
    if mapping is None:
        mapping = _build_mapping()
    brands.update(_normalize_brand(e[2]) for e in mapping if len(e) >= 3 and e[2])
    for name in os.listdir('.'):
        if not os.path.isdir(name):
            continue
        low = name.lower()
        if (_is_internal_dir(name) or low.startswith('backup')
                or name.startswith(INCREMENTAL_DIR_PREFIX)):
            continue
        if any(f.endswith('.xlsx') for _, _, fs in os.walk(name) for f in fs):
            brands.add(name)
    return sorted(brands)

def _find_mapping_entry(name_no_ext: str, mapping: list):
    """精确匹配：编号 + 语种均相符。"""
    base      = re.sub(TIMESTAMP_REGEX, '', name_no_ext)
    file_lang = _file_lang(name_no_ext)
    for entry in mapping:
        core_id, task_lang, *_ = entry
        if core_id in base and task_lang == file_lang:
            return entry
    return None

def _find_mapping_entry_with_fallback(name_no_ext: str, mapping: list):
    """
    先精确匹配；失败时尝试推断：
      - 编号相同 → 继承同一品牌
      - 语种后缀 → 从已有排期数据反推目录名
    返回 (entry, is_inferred)
    """
    entry = _find_mapping_entry(name_no_ext, mapping)
    if entry:
        return entry, False

    base      = re.sub(TIMESTAMP_REGEX, '', name_no_ext)
    file_lang = _file_lang(name_no_ext)
    nm        = re.search(r'\d{4,}', base)
    if not nm or not file_lang:
        return None, False

    core_id = nm.group()
    # 从排期中找同编号的任意条目 → 取品牌
    brand = next((e[2] for e in mapping if e[0] == core_id), None)
    # 从排期中找同语种后缀的任意条目 → 取目录名
    lang_dir = next((e[3] for e in mapping if e[1] == file_lang), None)

    if brand and lang_dir:
        inferred = (core_id, file_lang, brand, lang_dir, base)
        return inferred, True
    return None, False

def _collect_source_files() -> list:
    """
    收集待处理文件：优先读 _inbox/，inbox 为空时读根目录。
    返回 [(src_dir, filename), ...]
    """
    def _is_xlsx(f):
        return (f.endswith('.xlsx')
                and not f.startswith('~$')
                and f != SCHEDULE_FILE
                and 'backup' not in f.lower())

    # inbox 优先
    if os.path.isdir(INBOX_DIR):
        files = [(INBOX_DIR, f) for f in os.listdir(INBOX_DIR) if _is_xlsx(f)]
        if files:
            return files

    # 兜底：根目录
    brands = _get_brands()
    files = [('.', f) for f in os.listdir('.') if _is_xlsx(f)
             and not any(os.path.join('.', f).startswith(b) for b in brands)]
    return files

def _get_sent_samples(path: str, max_samples: int = 20) -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = next((wb[s] for s in wb.sheetnames if s.endswith('_sent')), None)
    if not ws:
        wb.close()
        return []
    out = []
    for row in ws.iter_rows(min_row=1, max_row=80, max_col=5, values_only=True):
        for v in row:
            s = str(v or '').strip()
            if len(s) > 3:
                out.append(s)
            if len(out) >= max_samples:
                wb.close()
                return out
    wb.close()
    return out

def _unique_path(path: str) -> str:
    """返回不覆盖已有文件的路径。"""
    if not os.path.exists(path):
        return path
    root, ext = os.path.splitext(path)
    i = 2
    while True:
        candidate = f"{root}_{i}{ext}"
        if not os.path.exists(candidate):
            return candidate
        i += 1


def _ensure_dir(path: str) -> str:
    os.makedirs(path, exist_ok=True)
    return path

def _is_internal_dir(name: str) -> bool:
    """内部目录（备份/输出/报告/归档、构建产物等），扫描车厂语料时应跳过。"""
    return (name.startswith('_') or name.startswith('.')
            or name in INTERNAL_DIRS
            or name in BUILD_ARTIFACT_DIRS
            or any(name.startswith(p) for p in BUILD_ARTIFACT_PREFIXES))

def _dest_exists_for(fname: str, mapping: list) -> bool:
    """判断源文件按排期归档后的目标是否已存在（用于跳过重复备份）。"""
    entry, _ = _find_mapping_entry_with_fallback(os.path.splitext(fname)[0], mapping)
    if not entry:
        return False
    _, _, brand, lang_dir, full_task_name = entry
    dest = os.path.join(brand, lang_dir, f"{full_task_name}_{TODAY_TIMESTAMP}.xlsx")
    return os.path.exists(dest)

def _archive_old(directory: str, pattern: str, keep: int = None,
                 kind: str = 'any', archive_sub: str = '') -> list:
    """把 directory 下匹配 pattern 的旧产物（保留最近 keep 份外）移动到 _archive/，不删除。"""
    keep = KEEP_RECENT if keep is None else keep
    if keep <= 0 or not os.path.isdir(directory):
        return []
    matches = sorted(glob.glob(os.path.join(directory, pattern)))
    if kind == 'dir':
        matches = [m for m in matches if os.path.isdir(m)]
    elif kind == 'file':
        matches = [m for m in matches if os.path.isfile(m)]
    old = matches[:-keep] if len(matches) > keep else []
    if not old:
        return []
    dest_root = _ensure_dir(os.path.join(ARCHIVE_DIR, archive_sub)
                            if archive_sub else ARCHIVE_DIR)
    moved = []
    for p in old:
        try:
            dest = _unique_path(os.path.join(dest_root, os.path.basename(p)))
            shutil.move(p, dest)
            moved.append(dest)
        except Exception as e:
            print(f"  [archive] 移动失败 {p}: {e}")
    return moved

def step8_prune(keep: int = None) -> None:
    """把各类历史产物移动到 _archive/（仅保留最近 keep 份），不删除任何数据。"""
    keep = KEEP_RECENT if keep is None else keep
    plan = [
        (BACKUPS_DIR, f"{BACKUP_DIR_PREFIX}_*",      'dir',  'backups'),
        (OUTPUTS_DIR, f"{INCREMENTAL_DIR_PREFIX}_*", 'dir',  'outputs'),
        (OUTPUTS_DIR, "new_corpus_*.zip",            'file', 'outputs'),
        (OUTPUTS_DIR, "final_new_only_*.zip",        'file', 'outputs'),
        (REPORTS_DIR, "处理报告_*.txt",              'file', 'reports'),
        (REPORTS_DIR, "排期_状态_*.xlsx",            'file', 'reports'),
        (TTS_TOOLS_DIR, "auto_tts_2*",               'dir',  'auto_tts'),
        (TTS_TOOLS_DIR, "auto_tts_2*.zip",           'file', 'auto_tts'),
    ]
    total = 0
    for directory, pattern, kind, sub in plan:
        total += len(_archive_old(directory, pattern, keep, kind, sub))
    if total:
        print(f"  [prune] 已将 {total} 项旧产物移动到 {ARCHIVE_DIR}/"
              f"（保留最近 {keep} 份，未删除）。")
    else:
        print(f"  [prune] 各类产物均未超过 {keep} 份，无需归档。")

def _parse_keep(argv: list, default: int = KEEP_RECENT) -> int:
    if '--keep' in argv:
        try:
            return int(argv[argv.index('--keep') + 1])
        except (IndexError, ValueError):
            pass
    return default


# ══════════════════════════════════════════════════════
# 核心处理：openpyxl 重命名 sheet，完整保留格式
# ══════════════════════════════════════════════════════

def process_excel(path: str) -> bool:
    name_no_ext = os.path.splitext(os.path.basename(path))[0]
    prefix      = re.sub(TIMESTAMP_REGEX, '', name_no_ext)
    sent_name   = f"{prefix}_sent"[:31]
    shuofa_name = f"{prefix}_shuofa"[:31]
    try:
        wb     = load_workbook(path)
        sheets = wb.sheetnames
        shuofa_cur = sent_cur = None

        # ── 第一轮：内容优先识别 ────────────────────────────
        # 先扫 header 判断 sheet 类型，不依赖 sheet 名
        for s in sheets:
            if s in ('<>', '-'):
                continue
            ws_tmp = wb[s]
            # 元数据表（功能需求/槽命名/验收集等）→ 直接跳过
            if _sheet_is_meta(ws_tmp):
                continue
            # 强关键词命中 shuofa，且内容通过验证
            if shuofa_cur is None and _is_shuofa_hard(s):
                if _shuofa_content_ok(ws_tmp):
                    shuofa_cur = s
                else:
                    print(f"  [警告] {os.path.basename(path)}: [{s}] 含大量中文，不视为 shuofa")
                continue
            # sent 关键词命中
            if sent_cur is None and _is_sent(s):
                sent_cur = s
                continue

        # ── 第二轮：名字兜底（内容无法判断的 sheet）──────────
        if shuofa_cur is None or sent_cur is None:
            for s in sheets:
                if s in ('<>', '-'):
                    continue
                if s in (shuofa_cur, sent_cur):
                    continue
                ws_tmp = wb[s]
                if _sheet_is_meta(ws_tmp):
                    continue
                # 已有 sent → 不再抢 sent 位；反之亦然
                if shuofa_cur is None and _is_shuofa_fallback(s) and _shuofa_content_ok(ws_tmp):
                    shuofa_cur = s

        if shuofa_cur is None and sent_cur is None:
            print(f"  [跳过] 无法识别 sheet 结构: {path}")
            return False

        # ── 重命名 shuofa ────────────────────────────────────
        if shuofa_cur and shuofa_cur != shuofa_name:
            wb[shuofa_cur].title = shuofa_name
            shuofa_cur = shuofa_name

        # ── 处理 sent ────────────────────────────────────────
        if sent_cur:
            ws = wb[sent_cur]
            # 用 _find_sent_col 定位逆规整目标列（比手写循环更清晰）
            target_col = _find_sent_col(ws)
            if target_col and ws.max_column > 1:
                for col in range(ws.max_column, 0, -1):
                    if col != target_col:
                        ws.delete_cols(col)
            # 去除中文 header 行
            rows_to_strip = 0
            for row in ws.iter_rows(min_row=1):
                if any(c.value and CHINESE_PATTERN.search(str(c.value)) for c in row):
                    rows_to_strip += 1
                else:
                    break
            if rows_to_strip:
                ws.delete_rows(1, rows_to_strip)
            if ws.title != sent_name:
                ws.title = sent_name
                sent_cur = sent_name
        # 删多余 sheet
        keep = set(filter(None, [shuofa_cur, sent_cur]))
        for s in ('<>', '-'):
            if s in wb.sheetnames:
                keep.add(s)
        for s in list(wb.sheetnames):
            if s not in keep:
                del wb[s]
        # 排序
        order = [shuofa_name, '<>', '-', sent_name]
        new_order = [wb[n] for n in order if n in wb.sheetnames]
        seen = {w.title for w in new_order}
        new_order += [w for w in wb.worksheets if w.title not in seen]
        wb._sheets = new_order
        wb.save(path)
        return True
    except Exception as e:
        print(f"  [错误] {path}: {e}")
        return False


# ══════════════════════════════════════════════════════
# Step 0：自动备份 inbox 原始文件
# ══════════════════════════════════════════════════════

def step0_backup(source_files: list, mapping: list) -> dict:
    """
    将待处理的原始文件备份到 _backups/backup_语料_时间戳/。
    已归档过（目标已存在）的文件跳过备份，避免重复堆积。
    """
    if not source_files:
        return {'dir': '', 'backed': 0, 'skipped': []}
    backup_dir = os.path.join(_ensure_dir(BACKUPS_DIR),
                              f"{BACKUP_DIR_PREFIX}_{RUN_TIMESTAMP}")
    backed, skipped = 0, []
    for src_dir, fname in source_files:
        if _dest_exists_for(fname, mapping):
            skipped.append(fname)
            continue
        os.makedirs(backup_dir, exist_ok=True)
        shutil.copy2(os.path.join(src_dir, fname),
                     os.path.join(backup_dir, fname))
        backed += 1
    skip_txt = f"（跳过已存在 {len(skipped)} 个）" if skipped else ""
    if backed:
        print(f"  [step0] 已备份 {backed} 个原始文件  ->  {backup_dir}/{skip_txt}")
    else:
        print(f"  [step0] 无需新增备份{skip_txt}")
    return {'dir': backup_dir if backed else '', 'backed': backed, 'skipped': skipped}


# ══════════════════════════════════════════════════════
# Step 1：整理文件到品牌/语种目录
# ══════════════════════════════════════════════════════

def step1_organize(source_files: list, mapping: list) -> dict:
    """
    将 source_files 按排期移动到品牌/语种目录。
    精确匹配失败时尝试推断（同编号继承品牌，语种后缀推断目录名）。
    """
    stats = {'moved': 0, 'moved_paths': [], 'inferred': [], 'unmatched': [],
             'unmatched_paths': [], 'skipped': 0, 'skipped_paths': []}
    if not mapping:
        print("  [step1] 排期.xlsx 不存在或为空，跳过整理。")
        return stats

    for src_dir, fname in source_files:
        src = os.path.join(src_dir, fname)
        entry, is_inferred = _find_mapping_entry_with_fallback(
            os.path.splitext(fname)[0], mapping)

        if entry and _cell_str(entry[2]) and _cell_str(entry[3]):
            _, _, brand, lang_dir, full_task_name = entry
            target_dir = os.path.join(brand, lang_dir)
            os.makedirs(target_dir, exist_ok=True)
            dest = os.path.join(target_dir,
                                f"{full_task_name}_{TODAY_TIMESTAMP}.xlsx")
            if os.path.exists(dest):
                duplicate_dir = os.path.join('待确认', '已存在')
                os.makedirs(duplicate_dir, exist_ok=True)
                duplicate_dest = _unique_path(os.path.join(duplicate_dir, fname))
                shutil.move(src, duplicate_dest)
                print(f"  [已存在] {fname}  ->  {duplicate_dest}")
                stats['skipped'] += 1
                stats['skipped_paths'].append(duplicate_dest)
            else:
                shutil.move(src, dest)
                if is_inferred:
                    print(f"  [推断移动] {fname}  ->  {dest}  <- 请在排期中补录！")
                    stats['inferred'].append(fname)
                else:
                    print(f"  [移动] {fname}  ->  {dest}")
                stats['moved'] += 1
                stats['moved_paths'].append(dest)
        else:
            misc = os.path.join('待确认', os.path.splitext(fname)[0])
            os.makedirs(misc, exist_ok=True)
            dest = _unique_path(os.path.join(misc, fname))
            shutil.move(src, dest)
            print(f"  [未匹配] {fname}  ->  待确认/")
            stats['unmatched'].append(fname)
            stats['unmatched_paths'].append(dest)

    print(f"  [step1] 移动 {stats['moved']}（推断 {len(stats['inferred'])}），"
          f"跳过 {stats['skipped']}，未匹配 {len(stats['unmatched'])}")
    return stats


# ══════════════════════════════════════════════════════
# Step 2：标准化 sheet 格式
# ══════════════════════════════════════════════════════

def _iter_all_corpus_files():
    for root, dirs, files in os.walk('.'):
        # 原地裁剪内部目录（_backups/_outputs/_reports/_archive/备份/增量等）
        dirs[:] = [d for d in dirs
                   if not _is_internal_dir(d)
                   and not d.lower().startswith('backup')
                   and not d.startswith(INCREMENTAL_DIR_PREFIX)]
        for f in files:
            if not f.endswith('.xlsx') or f == SCHEDULE_FILE or f.startswith('~$'):
                continue
            yield os.path.join(root, f)

def step2_cleanup(paths: list = None) -> dict:
    """
    标准化 Excel 格式。
    paths 为空时全量扫描；传入路径时仅处理本批新增文件，避免反复改写历史语料。
    """
    ok = fail = 0
    ok_paths, fail_paths = [], []
    targets = list(paths) if paths is not None else list(_iter_all_corpus_files())
    for path in targets:
        if process_excel(path):
            ok += 1
            ok_paths.append(path)
        else:
            fail += 1
            fail_paths.append(path)
    scope = '本批' if paths is not None else '全量'
    print(f"  [step2] {scope}标准化：成功 {ok} 个，失败/跳过 {fail} 个")
    return {'ok': ok, 'fail': fail, 'scope': scope, 'total': len(targets),
            'ok_paths': ok_paths, 'fail_paths': fail_paths}


# ══════════════════════════════════════════════════════
# Step 3：导出本批新增语料
# ══════════════════════════════════════════════════════

def step3_export_incremental(paths: list) -> dict:
    """
    将本批新增且标准化成功的文件复制到独立时间戳目录。
    目录保留 车厂/语种/文件.xlsx 结构，便于直接查看本次相对上次的新增语料。
    """
    paths = [p for p in paths if p and os.path.exists(p)]
    if not paths:
        print("  [step3] 本批无标准化成功的新文件，跳过新增语料目录。")
        return {'dir': '', 'count': 0}

    output_dir = os.path.join(_ensure_dir(OUTPUTS_DIR),
                              f"{INCREMENTAL_DIR_PREFIX}_{RUN_TIMESTAMP}")
    count = 0
    for path in paths:
        rel_path = os.path.relpath(path, '.')
        dest = os.path.join(output_dir, rel_path)
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        shutil.copy2(path, dest)
        count += 1

    print(f"  [step3] 已生成本批新增语料目录：{output_dir}/（{count} 个文件）")
    return {'dir': output_dir, 'count': count}


# ══════════════════════════════════════════════════════
# Step 4：自动合成新增语料
# ══════════════════════════════════════════════════════

def _infer_tts_lang(path: str) -> str:
    name_no_ext = os.path.splitext(os.path.basename(path))[0]
    file_lang = _file_lang(name_no_ext)
    if file_lang:
        return file_lang
    dir_tts_langs = {
        '阿语': 'ar_il',
        '德语': 'de',
        '法语': 'fr_fr',
        '希伯来语': 'he_il',
        '葡语': 'pt_la',
        '葡萄牙语': 'pt_la',
        '泰语': 'th_th',
        '丹麦语': 'da_dk',
        '挪威语': 'nb_no',
        '瑞典语': 'sv_se',
        '荷兰语': 'nl_nl',
        '匈牙利语': 'hu',
        '俄语': 'ru',
        '斯洛文尼亚语': 'si_si',
    }
    parts = os.path.normpath(path).split(os.sep)
    for part in reversed(parts[:-1]):
        lang = dir_tts_langs.get(part) or DIR_TTS_LANGS.get(part)
        if lang:
            return lang
    return ''

def _load_tts_config() -> dict:
    tool = os.path.join(TTS_TOOLS_DIR, 'asr_tts_tool.py')
    if not os.path.exists(tool):
        return {}
    old_path = list(sys.path)
    try:
        if TTS_TOOLS_DIR not in sys.path:
            sys.path.insert(0, TTS_TOOLS_DIR)
        import asr_tts_tool
        config = {}
        for _, (lang_id, voices) in asr_tts_tool.LANG_CONFIG.items():
            for voice, _gender in voices:
                locale = '-'.join(voice.split('-')[:2])
                config.setdefault(locale, (lang_id, voice))
        return config
    except Exception as e:
        print(f"  [tts] 读取 TTS 配置失败：{e}")
        return {}
    finally:
        sys.path = old_path

def _extract_sent_lines(path: str) -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = next((wb[s] for s in wb.sheetnames if s.endswith('_sent')), None)
        if not ws:
            return []
        lines = []
        for row in ws.iter_rows(values_only=True):
            text = ''
            for v in row:
                text = str(v or '').strip()
                if text:
                    break
            if text:
                lines.append(text)
        return lines
    finally:
        wb.close()

def step4_synthesize_incremental(incremental_dir: str) -> dict:
    """
    对本次新增目录中的 Excel 自动合成测试集。
    每个 Excel 最多随机取 1000 条 _sent，不足则全取；测试集目录按文件名命名。
    """
    result = {'enabled': False, 'output_dir': '', 'zip': '', 'ok': [],
              'skipped': [], 'failed': []}
    if not incremental_dir or not os.path.isdir(incremental_dir):
        print("  [step4] 无新增语料目录，跳过自动合成。")
        return result

    tool = os.path.join(TTS_TOOLS_DIR, 'asr_tts_tool.py')
    if not os.path.exists(tool):
        print(f"  [step4] 未找到 TTS 工具：{tool}，跳过自动合成。")
        result['skipped'].append(f"missing tool: {tool}")
        return result

    tts_config = _load_tts_config()
    if not tts_config:
        print("  [step4] TTS 语言配置为空，跳过自动合成。")
        result['skipped'].append("empty tts config")
        return result

    output_dir = os.path.join(TTS_TOOLS_DIR, f"auto_tts_{RUN_TIMESTAMP}")
    os.makedirs(output_dir, exist_ok=True)
    result['enabled'] = True
    result['output_dir'] = output_dir

    excel_files = sorted(
        os.path.join(root, f)
        for root, _, files in os.walk(incremental_dir)
        for f in files
        if f.endswith('.xlsx') and not f.startswith('~$')
    )
    if not excel_files:
        print("  [step4] 新增目录中没有 Excel，跳过自动合成。")
        return result

    # 云端/本地同名去重：同名语料只合成本地那份测试集
    #   规则1 同号仅 -L-/-C- 不同；规则2 NS 单编号差±1 且排期中 车厂+语种 相同（双保险）
    def _norm_lc(p):
        return re.sub(r'-(?:L|C)-ASR', '-ASR', os.path.basename(p))

    def _core(p):
        m = re.search(r'(\d{4,})', os.path.basename(p))
        return m.group(1) if m else None

    _mapping = _build_mapping()

    def _sched_bl(core):
        e = next((e for e in _mapping if e[0] == core), None)
        return (e[2], e[3]) if e else None

    _locals = {_norm_lc(p) for p in excel_files if '-L-ASR' in os.path.basename(p)}
    _locals_ns = []
    for p in excel_files:
        name = os.path.basename(p)
        if '-L-ASR' in name and name.startswith('NS-'):
            c = _core(p)
            if c:
                _locals_ns.append((int(c), _sched_bl(c)))

    deduped = []
    for p in excel_files:
        name = os.path.basename(p)
        dup = False
        if '-C-ASR' in name:
            if _norm_lc(p) in _locals:
                dup = True
            elif name.startswith('NS-'):
                c = _core(p)
                bl = _sched_bl(c) if c else None
                if c and bl:
                    dup = any(n2 is not None and abs(int(c) - n2) == 1 and bl2 == bl
                              for n2, bl2 in _locals_ns)
        if dup:
            msg = f"{os.path.splitext(name)[0]}: 与本地同名，测试集共用（跳过云端）"
            print(f"    [跳过] {msg}")
            result['skipped'].append(msg)
        else:
            deduped.append(p)
    excel_files = deduped

    print(f"  [step4] 自动合成 {len(excel_files)} 个新增 Excel")
    for path in excel_files:
        test_name = os.path.splitext(os.path.basename(path))[0]
        lang = _infer_tts_lang(path)
        locale = TTS_LANG_LOCALES.get(lang)
        if not locale:
            msg = f"{test_name}: 未配置 TTS 语言映射（{lang or 'unknown'}）"
            print(f"    [跳过] {msg}")
            result['skipped'].append(msg)
            continue
        entry = tts_config.get(locale)
        if not entry:
            msg = f"{test_name}: TTS 工具无 voice（{locale}）"
            print(f"    [跳过] {msg}")
            result['skipped'].append(msg)
            continue

        lines = _extract_sent_lines(path)
        if not lines:
            msg = f"{test_name}: 无 _sent 内容"
            print(f"    [跳过] {msg}")
            result['skipped'].append(msg)
            continue
        selected = random.sample(lines, TTS_SAMPLE_SIZE) if len(lines) > TTS_SAMPLE_SIZE else lines

        test_dir = _unique_path(os.path.join(output_dir, test_name))
        os.makedirs(test_dir, exist_ok=True)
        input_path = os.path.join(test_dir, 'input.txt')
        with open(input_path, 'w', encoding='utf-8') as fp:
            fp.write('\n'.join(selected))
            fp.write('\n')

        lang_id, voice = entry
        cmd = [
            sys.executable, tool, '--cli',
            '-i', input_path,
            '-d', test_dir,
            '-l', str(lang_id),
            '-v', voice,
            '-c', str(TTS_CONCURRENCY),
            '--spd-ratio', '0',
        ]
        print(f"    [合成] {test_name}  {len(selected)}/{len(lines)}  {locale}  {voice}", flush=True)
        try:
            log_path = os.path.join(test_dir, 'tts.log')
            with open(log_path, 'w', encoding='utf-8') as logf:
                proc = subprocess.Popen(
                    cmd, cwd=TTS_TOOLS_DIR, text=True,
                    stdout=logf, stderr=subprocess.STDOUT,
                    encoding='utf-8', errors='replace')
                # 边合成边报进度：数 test_dir 里已生成的 wav
                last = -1
                while proc.poll() is None:
                    time.sleep(2)
                    try:
                        n = len([f for f in os.listdir(test_dir)
                                 if f.lower().endswith('.wav')])
                    except OSError:
                        n = last
                    if n != last:
                        print(f"    [{n}/{len(selected)}] 合成中 {test_name}", flush=True)
                        last = n
            if proc.returncode != 0:
                raise RuntimeError(f"TTS exited {proc.returncode}")
            wav_count = len([f for f in os.listdir(test_dir) if f.lower().endswith('.wav')])
            test_zip_path = test_dir + '.zip'
            if os.path.exists(test_zip_path):
                os.remove(test_zip_path)
            with zipfile.ZipFile(test_zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
                for root, _, files in os.walk(test_dir):
                    for f in files:
                        fpath = os.path.join(root, f)
                        rel = os.path.relpath(fpath, test_dir)
                        z.write(fpath, os.path.join(test_name, rel))
            result['ok'].append({
                'name': test_name, 'selected': len(selected), 'total': len(lines),
                'wav': wav_count, 'dir': test_dir, 'locale': locale, 'voice': voice,
                'zip': test_zip_path,
            })
        except Exception as e:
            msg = f"{test_name}: {e}"
            print(f"    [失败] {msg}")
            result['failed'].append(msg)

    summary_path = os.path.join(output_dir, 'synthesis_summary.tsv')
    with open(summary_path, 'w', encoding='utf-8') as fp:
        fp.write("name\tselected\ttotal\twav\tlocale\tvoice\tdir\tzip\n")
        for item in result['ok']:
            fp.write(
                f"{item['name']}\t{item['selected']}\t{item['total']}\t"
                f"{item['wav']}\t{item['locale']}\t{item['voice']}\t"
                f"{item['dir']}\t{item['zip']}\n")

    zip_path = output_dir + '.zip'
    if os.path.exists(zip_path):
        os.remove(zip_path)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
        z.write(summary_path, os.path.basename(summary_path))
        for item in result['ok']:
            z.write(item['zip'], os.path.basename(item['zip']))
    result['zip'] = zip_path
    print(f"  [step4] 合成完成：成功 {len(result['ok'])}，跳过 {len(result['skipped'])}，失败 {len(result['failed'])}")
    print(f"  [step4] 输出：{output_dir}")
    print(f"  [step4] ZIP：{zip_path}")
    return result


# ══════════════════════════════════════════════════════
# Step 5：打包 ZIP
# ══════════════════════════════════════════════════════

def step3_zip(output: str = None, mapping: list = None) -> dict:
    output = output or os.path.join(_ensure_dir(OUTPUTS_DIR), 'processed_corpus.zip')
    if os.path.exists(output):
        os.remove(output)
    count = 0
    brands = _get_brands(mapping)
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as z:
        for b in brands:
            if not os.path.exists(b):
                continue
            for root, _, fs in os.walk(b):
                for f in fs:
                    fpath = os.path.join(root, f)
                    z.write(fpath)
                    count += 1
    size_mb = os.path.getsize(output) / 1024 / 1024
    print(f"  [step3] 已打包 {count} 个文件，{size_mb:.1f} MB  ->  {output}")
    return {'count': count, 'size_mb': size_mb, 'output': output,
            'brands': brands}


def _zip_ascii_part(part: str) -> str:
    return ZIP_BRAND_NAMES.get(part) or ZIP_LANG_NAMES.get(part) or part


def _zip_incremental_corpus_ascii(incremental_dir: str) -> str:
    if not incremental_dir or not os.path.isdir(incremental_dir):
        return ''
    zip_path = os.path.join(_ensure_dir(OUTPUTS_DIR), f"new_corpus_{RUN_TIMESTAMP}.zip")
    if os.path.exists(zip_path):
        os.remove(zip_path)
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(incremental_dir):
            for f in files:
                if not f.endswith('.xlsx') or f.startswith('~$'):
                    continue
                fpath = os.path.join(root, f)
                rel = os.path.relpath(fpath, incremental_dir)
                arcname = '/'.join(_zip_ascii_part(p) for p in rel.split(os.sep))
                z.write(fpath, arcname)
    return zip_path


def step5_final_package(incremental_dir: str, tts_zip: str = '', report_path: str = '',
                        output: str = None, schedule_status: str = '') -> dict:
    """Package only this run's new corpus, synthesized test sets, and report into one ZIP."""
    output = output or os.path.join(_ensure_dir(OUTPUTS_DIR),
                                    f"final_new_only_{RUN_TIMESTAMP}.zip")
    if os.path.exists(output):
        os.remove(output)

    new_corpus_zip = _zip_incremental_corpus_ascii(incremental_dir)
    items = []
    for src, arc in (
        (new_corpus_zip, os.path.basename(new_corpus_zip) if new_corpus_zip else ''),
        (tts_zip, f"tts_test_sets_{RUN_TIMESTAMP}.zip" if tts_zip else ''),
        (report_path, f"report_{RUN_TIMESTAMP}.txt" if report_path else ''),
        (schedule_status,
         f"schedule_status_{RUN_TIMESTAMP}.xlsx" if schedule_status else ''),
    ):
        if src and os.path.exists(src):
            items.append((src, arc))

    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as z:
        for src, arc in items:
            z.write(src, arc)

    size_mb = os.path.getsize(output) / 1024 / 1024
    print(f"  [final] 已打包 {len(items)} 个文件，{size_mb:.1f} MB  ->  {output}")
    return {'output': output, 'count': len(items), 'size_mb': size_mb,
            'items': [src for src, _ in items]}


# ══════════════════════════════════════════════════════
# Step 9：跨网段传输（rftctl local submit）
# ══════════════════════════════════════════════════════

def _find_rftctl() -> str:
    """定位 rftctl：环境变量 > PATH > 默认安装路径。"""
    p = os.environ.get('RFTCTL_PATH')
    if p and os.path.exists(p):
        return p
    found = shutil.which('rftctl')
    if found:
        return found
    if os.path.exists(RFTCTL_PATH_DEFAULT):
        return RFTCTL_PATH_DEFAULT
    return ''

def _parse_receives(spec) -> list:
    """接收网段：list 或逗号分隔字符串 -> 去空去重（保序）的列表。"""
    items = spec if isinstance(spec, (list, tuple)) else str(spec).split(',')
    out = []
    for r in items:
        r = str(r).strip()
        if r and r not in out:
            out.append(r)
    return out

def step9_transfer(files: list, receives=None) -> dict:
    """
    把本批产物（new_corpus + TTS 测试集）通过 rftctl 上传到一个或多个接收网段。
    同一份文件只硬链接暂存一次，对每个网段各发起一个传输流程；全部成功才清理暂存夹。
    """
    recvs = _parse_receives(receives if receives is not None else TRANSFER_RECEIVES)
    result = {'receives': recvs, 'sent': [], 'failed': [], 'files': [], 'error': ''}
    files = [f for f in files if f and os.path.exists(f)]
    if not files:
        print("  [step9] 无可传输文件，跳过。")
        return result
    if not recvs:
        print("  [step9] 未指定接收网段，跳过。")
        return result

    rft = _find_rftctl()
    if not rft:
        result['error'] = 'rftctl not found'
        print("  [step9] 未找到 rftctl（设置 RFTCTL_PATH 或加入 PATH），跳过传输。")
        return result

    stage = os.path.join(_ensure_dir(OUTPUTS_DIR), f"_transfer_{RUN_TIMESTAMP}")
    if os.path.exists(stage):
        shutil.rmtree(stage)
    os.makedirs(stage)
    for f in files:
        dst = os.path.join(stage, os.path.basename(f))
        try:
            os.link(f, dst)           # 同盘硬链接，秒级、不占额外空间
        except OSError:
            shutil.copy2(f, dst)      # 跨盘退回复制
        result['files'].append(os.path.basename(f))

    remark = f"{TODAY_TIMESTAMP}传输语料"
    title  = f"ASR语料测试集{TODAY_TIMESTAMP}"
    print(f"  [step9] 传输 {len(files)} 个文件到 {recvs}：{', '.join(result['files'])}")
    for recv in recvs:
        cmd = [rft, 'local', 'submit', '--receive', recv,
               '--file', stage, '--remark', remark, '--title', title]
        cmd_str = ' '.join(f'"{c}"' if ' ' in c else c for c in cmd)
        print(f"  [step9] -> [{recv}] 提交中…")
        try:
            completed = subprocess.run(cmd, text=True, encoding='utf-8', errors='replace')
            if completed.returncode == 0:
                result['sent'].append(recv)
                print(f"  [step9] [{recv}] 传输完成。")
            else:
                result['failed'].append(recv)
                print(f"  [step9] [{recv}] 传输失败（exit {completed.returncode}）。重试：{cmd_str}")
        except Exception as e:
            result['failed'].append(recv)
            print(f"  [step9] [{recv}] 传输异常：{e}")

    if result['failed']:
        print(f"  [step9] 有失败网段 {result['failed']}，暂存夹保留：{stage}")
    else:
        shutil.rmtree(stage, ignore_errors=True)   # 全部成功才清理（硬链接，不动原件）
    print(f"  [step9] 汇总：成功 {result['sent']}，失败 {result['failed']}")
    return result

def _latest_run_products() -> list:
    """找最近一次运行的 new_corpus 与配套 auto_tts 产物，供 send 子命令重发。"""
    ncs = sorted(glob.glob(os.path.join(OUTPUTS_DIR, "new_corpus_*.zip")))
    ttss = sorted(glob.glob(os.path.join(TTS_TOOLS_DIR, "auto_tts_2*.zip")))
    files = []
    if ncs:
        nc = ncs[-1]
        files.append(nc)
        # 用同一 run 时间戳找配套的 auto_tts，找不到再退回最新
        m = re.search(r'new_corpus_(\d{8}_\d{6})\.zip$', nc)
        paired = (os.path.join(TTS_TOOLS_DIR, f"auto_tts_{m.group(1)}.zip")
                  if m else '')
        if paired and os.path.exists(paired):
            files.append(paired)
        elif ttss:
            files.append(ttss[-1])
    elif ttss:
        files.append(ttss[-1])
    return files


# ══════════════════════════════════════════════════════
# Step 6：三重核验
# ══════════════════════════════════════════════════════

def step4_verify(mapping: list) -> dict:
    """
    核验 1：排期文件完整性
    核验 2：sheet 结构（_sent / _shuofa）
    核验 3：内容与语种匹配
    返回结果 dict，同时写到报告中。
    """
    result = {'pass': True, 'missing': [], 'struct_errors': [],
              'struct_warns': [], 'lang_errors': [], 'lang_ok': 0,
              'lang_skip': 0, 'file_list': [], 'distribution': {}}

    # ── 核验 1：文件完整性 ──────────────────────────────
    print("  -- 核验 1/3：文件完整性（排期对照）")
    for entry in mapping:
        core_id, task_lang, brand, lang_dir, full_task_name = entry
        target_dir = os.path.join(brand, lang_dir)
        found = False
        if os.path.exists(target_dir):
            for fname in os.listdir(target_dir):
                if not fname.endswith('.xlsx'):
                    continue
                matched = _find_mapping_entry(os.path.splitext(fname)[0], mapping)
                if matched and matched[4] == full_task_name:
                    found = True
                    break
        if not found:
            result['missing'].append(f"{full_task_name}  ({brand}/{lang_dir})")

    if result['missing']:
        result['pass'] = False
        print(f"  [FAIL] 缺少 {len(result['missing'])} 个任务文件：")
        for t in result['missing']:
            print(f"    [X] {t}")
    else:
        print(f"  [PASS] 排期中 {len(mapping)} 个任务全部有对应文件")

    # ── 核验 2：sheet 结构 ──────────────────────────────
    print("  -- 核验 2/3：Sheet 结构")
    for b in _get_brands(mapping):
        if not os.path.exists(b):
            continue
        for root, _, files in os.walk(b):
            for f in files:
                if not f.endswith('.xlsx'):
                    continue
                path = os.path.join(root, f)
                result['file_list'].append((path, root))
                # 统计分布
                key = os.path.join(*os.path.normpath(root).split(os.sep)[:2])
                result['distribution'].setdefault(key, []).append(f)
                try:
                    wb = load_workbook(path, read_only=True)
                    sheets = wb.sheetnames
                    wb.close()
                    if not any(s.endswith('_sent') for s in sheets):
                        result['struct_errors'].append(f"无 _sent: {path}")
                    elif not any(s.endswith('_shuofa') for s in sheets):
                        result['struct_warns'].append(f"无 _shuofa（原始格式）: {path}")
                except Exception as e:
                    result['struct_errors'].append(f"读取失败: {path}: {e}")

    hard = result['struct_errors']
    if hard:
        result['pass'] = False
        print(f"  [FAIL] {len(hard)} 个文件结构异常：")
        for e in hard: print(f"    [X] {e}")
    else:
        print(f"  [PASS] {len(result['file_list'])} 个文件均含 _sent sheet")
    if result['struct_warns']:
        print(f"  [WARN] {len(result['struct_warns'])} 个仅有 sent（原始格式，无 shuofa）")

    # ── 核验 3：内容与语种匹配 ──────────────────────────
    print("  -- 核验 3/3：内容与语种匹配")
    for path, root in result['file_list']:
        name_no_ext = os.path.splitext(os.path.basename(path))[0]
        file_lang   = _file_lang(name_no_ext)
        validator   = LANG_VALIDATORS.get(file_lang)
        if validator is None:
            dir_parts = os.path.normpath(root).split(os.sep)
            validator = DIR_LANG_VALIDATORS.get(dir_parts[-1] if dir_parts else '')
        if validator is None:
            result['lang_skip'] += 1
            continue
        samples = _get_sent_samples(path)
        if not samples:
            result['lang_errors'].append(f"sent 无内容: {path}")
            continue
        ratio = sum(1 for s in samples if validator(s)) / len(samples)
        if ratio < 0.5:
            result['lang_errors'].append(
                f"语种不符（命中 {ratio:.0%}）: {path}  样本: {samples[:1]}")
            result['pass'] = False
        else:
            result['lang_ok'] += 1

    if result['lang_errors']:
        result['pass'] = False
        print(f"  [FAIL] {len(result['lang_errors'])} 个语种不符：")
        for e in result['lang_errors']: print(f"    [X] {e}")
    else:
        print(f"  [PASS] {result['lang_ok']} 个语种正确，"
              f"{result['lang_skip']} 个无法自动判断（跳过）")

    print()
    if result['pass']:
        print("  [OK] 全部核验通过，数据完整可信。")
    else:
        print("  [FAIL] 存在问题，请核查上方错误。")
    return result


# ══════════════════════════════════════════════════════
# 生成处理报告
# ══════════════════════════════════════════════════════

def write_report(stats: dict, report_path: str):
    lines = []
    W = 52
    lines.append("处理报告  " + stats['run_ts'])
    lines.append("=" * W)

    lines.append(f"\n来源目录  : {stats.get('source_dir', '根目录')}"
                 f"  （{stats.get('source_count', 0)} 个文件）")
    s0 = stats.get('step0', {})
    if stats.get('backup_dir'):
        skip_txt = (f"，跳过已存在 {len(s0['skipped'])} 个"
                    if s0.get('skipped') else "")
        lines.append(f"自动备份  : {stats['backup_dir']}/"
                     f"  （新增备份 {s0.get('backed', 0)} 个{skip_txt}）")
    elif s0.get('skipped'):
        lines.append(f"自动备份  : 无新增（全部为已存在，跳过 {len(s0['skipped'])} 个）")

    lines.append(f"\n── Step 1 整理 {'─'*(W-12)}")
    s1 = stats.get('step1', {})
    lines.append(f"  精确匹配并移动 : {s1.get('moved', 0) - len(s1.get('inferred', []))} 个")
    if s1.get('inferred'):
        lines.append(f"  推断匹配并移动 : {len(s1['inferred'])} 个  ← 请在排期中补录！")
        for f in s1['inferred']:
            lines.append(f"    · {f}")
    lines.append(f"  已存在跳过     : {s1.get('skipped', 0)} 个")
    if s1.get('unmatched'):
        lines.append(f"  未匹配（待确认）: {len(s1['unmatched'])} 个")
        for f in s1['unmatched']:
            lines.append(f"    · {f}")

    lines.append(f"\n── Step 2 格式 {'─'*(W-12)}")
    s2 = stats.get('step2', {})
    lines.append(f"  范围 {s2.get('scope','全量')}，共 {s2.get('total',0)} 个")
    lines.append(f"  成功 {s2.get('ok',0)} 个，失败/跳过 {s2.get('fail',0)} 个")

    lines.append(f"\n── Step 3 新增 {'─'*(W-12)}")
    inc = stats.get('step3_incremental', {})
    if inc.get('dir'):
        lines.append(f"  {inc['dir']}/  {inc.get('count',0)} 个文件")
    else:
        lines.append("  本批无新增语料目录")

    lines.append(f"\n── Step 4 合成 {'─'*(W-12)}")
    tts = stats.get('step4_tts', {})
    if tts.get('enabled'):
        lines.append(f"  输出目录: {tts.get('output_dir','')}")
        lines.append(f"  ZIP     : {tts.get('zip','')}")
        lines.append(f"  成功 {len(tts.get('ok',[]))} 个，跳过 {len(tts.get('skipped',[]))} 个，失败 {len(tts.get('failed',[]))} 个")
        for item in tts.get('ok', []):
            zip_name = os.path.basename(item.get('zip', ''))
            zip_part = f"  zip {zip_name}" if zip_name else ""
            lines.append(f"    · {item['name']}  wav {item['wav']}/{item['selected']}  {item['locale']}{zip_part}")
        for item in tts.get('skipped', []):
            lines.append(f"    · 跳过: {item}")
        for item in tts.get('failed', []):
            lines.append(f"    · 失败: {item}")
    else:
        lines.append("  未执行自动合成")

    lines.append(f"\n── Step 5 打包 {'─'*(W-12)}")
    s3 = stats.get('step3', {})
    lines.append(f"  {s3.get('output','processed_corpus.zip')}"
                 f"  {s3.get('count',0)} 个文件  {s3.get('size_mb',0):.1f} MB")
    if s3.get('brands'):
        lines.append(f"  车厂范围: {', '.join(s3['brands'])}")

    lines.append(f"\n── Step 6 核验 {'─'*(W-12)}")
    v = stats.get('verify', {})
    lines.append(f"  文件完整性 : {'通过' if not v.get('missing') else '失败 ' + str(len(v['missing'])) + ' 个缺失'}")
    lines.append(f"  Sheet结构  : {'通过' if not v.get('struct_errors') else '失败 ' + str(len(v['struct_errors'])) + ' 个异常'}")
    lines.append(f"  语种匹配   : {'通过' if not v.get('lang_errors') else '失败 ' + str(len(v['lang_errors'])) + ' 个不符'}")

    dist = v.get('distribution', {})
    if dist:
        lines.append(f"\n── 文件分布 {'─'*(W-12)}")
        for key in sorted(dist):
            files = dist[key]
            lines.append(f"  {key}:  {len(files)} 个")
            for f in sorted(files):
                lines.append(f"    · {f}")

    if stats.get('schedule_status'):
        lines.append(f"\n── 排期状态 {'─'*(W-11)}")
        lines.append(f"  {stats['schedule_status']}  (状态列: 已存在/本次新增/缺失)")

    lines.append("\n" + "=" * W)
    lines.append("总结: " + ("✔ 全部通过，数据完整可信。"
                             if v.get('pass') else "✘ 存在问题，请核查报告。"))

    with open(report_path, 'w', encoding='utf-8') as fp:
        fp.write('\n'.join(lines))
    print(f"\n  已写入处理报告  ->  {report_path}")


# ══════════════════════════════════════════════════════
# 排期状态回写
# ══════════════════════════════════════════════════════

def _scan_archived_tasks(mapping: list) -> dict:
    """扫描车厂目录，返回 {排期任务名: 归档路径}。"""
    present = {}
    for b in _get_brands(mapping):
        if not os.path.exists(b):
            continue
        for root, _, files in os.walk(b):
            for f in files:
                if not f.endswith('.xlsx') or f.startswith('~$'):
                    continue
                matched = _find_mapping_entry(os.path.splitext(f)[0], mapping)
                if matched:
                    present.setdefault(matched[4], os.path.join(root, f))
    return present

def write_schedule_status(mapping: list, moved_names: set, out_path: str) -> str:
    """
    在排期旁生成带『状态』列的副本：已存在 / 本次新增 / 缺失。
    不改动原排期，359 行里一眼看清哪些已归档、哪些还缺。
    """
    if not os.path.exists(SCHEDULE_FILE):
        return ''
    df = pd.read_excel(SCHEDULE_FILE)
    present = _scan_archived_tasks(mapping)
    statuses, paths = [], []
    for _, row in df.iterrows():
        task_id = str(row.iloc[0]).strip()
        entry = _find_mapping_entry(task_id, mapping)
        full = entry[4] if entry else task_id
        p = present.get(full)
        if p:
            statuses.append('本次新增' if full in moved_names else '已存在')
            paths.append(os.path.relpath(p, '.'))
        else:
            statuses.append('缺失')
            paths.append('')
    df['状态'] = statuses
    df['归档路径'] = paths
    df.to_excel(out_path, index=False)
    n_have = sum(1 for s in statuses if s != '缺失')
    print(f"  已写入排期状态  ->  {out_path}  （已归档 {n_have}/{len(statuses)}）")
    return out_path


# ══════════════════════════════════════════════════════
# 恢复备份
# ══════════════════════════════════════════════════════

def restore_from_backup(backup_dir: str = None):
    if backup_dir is None:
        search_roots = [BACKUPS_DIR, '.', os.path.join(ARCHIVE_DIR, 'backups')]
        candidates = []
        for r in search_roots:
            if not os.path.isdir(r):
                continue
            for d in os.listdir(r):
                full = os.path.join(r, d)
                if d.lower().startswith('backup') and os.path.isdir(full):
                    candidates.append(full)
        if not candidates:
            print("未找到备份目录。")
            return
        backup_dir = sorted(candidates, key=os.path.basename)[-1]
    if not os.path.isdir(backup_dir):
        print(f"备份目录不存在: {backup_dir}")
        return

    mapping = _build_mapping()
    print(f"正在从备份恢复: {backup_dir}")
    restored = 0
    unmatched = []
    for f in sorted(os.listdir(backup_dir)):
        if not f.endswith('.xlsx'):
            continue
        entry, _ = _find_mapping_entry_with_fallback(
            os.path.splitext(f)[0], mapping)
        if entry and _cell_str(entry[2]) and _cell_str(entry[3]):
            _, _, brand, lang_dir, full_task_name = entry
            target_dir = os.path.join(brand, lang_dir)
            os.makedirs(target_dir, exist_ok=True)
            dest = os.path.join(target_dir, f"{full_task_name}_{TODAY_TIMESTAMP}.xlsx")
            shutil.copy2(os.path.join(backup_dir, f), dest)
            print(f"  [恢复] {f}  ->  {dest}")
            restored += 1
        else:
            unmatched.append(f)

    print(f"\n恢复完成: {restored} 个，未匹配 {len(unmatched)} 个")
    if unmatched:
        print("  未匹配：", ', '.join(unmatched))

    mapping = _build_mapping()
    print("\n正在标准化已恢复文件…")
    step2_cleanup()
    print("\n=== 恢复后核验 ===")
    step4_verify(mapping)


# ══════════════════════════════════════════════════════
# 主入口
# ══════════════════════════════════════════════════════

if __name__ == "__main__":
    if len(sys.argv) > 1:
        cmd = sys.argv[1]
        if cmd == 'restore':
            restore_from_backup(sys.argv[2] if len(sys.argv) > 2 else None)
            sys.exit(0)
        if cmd == 'verify':
            ok = step4_verify(_build_mapping())
            sys.exit(0 if ok['pass'] else 1)
        if cmd == 'clean':
            result = step2_cleanup()
            sys.exit(0 if result['fail'] == 0 else 1)
        if cmd == 'package':
            step3_zip(mapping=_build_mapping())
            sys.exit(0)
        if cmd == 'prune':
            step8_prune(_parse_keep(sys.argv))
            sys.exit(0)
        if cmd == 'status':
            _build = _build_mapping()
            write_schedule_status(
                _build, set(),
                os.path.join(_ensure_dir(REPORTS_DIR),
                             f"排期_状态_{RUN_TIMESTAMP}.xlsx"))
            sys.exit(0)
        if cmd == 'send':
            if '--file' in sys.argv:
                send_files = [sys.argv[sys.argv.index('--file') + 1]]
            else:
                send_files = _latest_run_products()
            recv = (sys.argv[sys.argv.index('--receive') + 1]
                    if '--receive' in sys.argv else TRANSFER_RECEIVES)
            res = step9_transfer(send_files, recv)
            sys.exit(0 if res['sent'] and not res['failed'] else 1)

    # ── 初始化 _inbox/ ──────────────────────────────
    os.makedirs(INBOX_DIR, exist_ok=True)

    source_files = _collect_source_files()
    source_dir   = INBOX_DIR if source_files and source_files[0][0] == INBOX_DIR else '根目录'
    mapping      = _build_mapping()

    # ── 语种/车厂筛选（--lang/--brand 或环境变量 FILTER_LANG/FILTER_BRAND；留空=全量）──
    def _arg_val(flag):
        if flag in sys.argv:
            i = sys.argv.index(flag)
            if i + 1 < len(sys.argv):
                return sys.argv[i + 1]
        return ''
    _flt_lang  = (_arg_val('--lang')  or os.environ.get('FILTER_LANG', '')).strip()
    _flt_brand = (_arg_val('--brand') or os.environ.get('FILTER_BRAND', '')).strip()
    if _flt_lang or _flt_brand:
        def _match_flt(sf):
            entry, _ = _find_mapping_entry_with_fallback(os.path.splitext(sf[1])[0], mapping)
            if not entry:
                return False
            b, l = _cell_str(entry[2]), _cell_str(entry[3])
            if _flt_lang and _flt_lang not in l:
                return False
            if _flt_brand and _flt_brand not in b:
                return False
            return True
        _before = len(source_files)
        source_files = [sf for sf in source_files if _match_flt(sf)]
        print(f"=== 筛选：语种[{_flt_lang or '不限'}] 车厂[{_flt_brand or '不限'}]："
              f"{_before} -> {len(source_files)} 个文件 ===")

    stats        = {'run_ts': RUN_TIMESTAMP, 'source_dir': source_dir,
                    'source_count': len(source_files)}

    if not source_files:
        print(f"_inbox/ 为空，没有新文件需要处理。")
        print("请将待处理的 xlsx 文件放入 _inbox/ 目录后重新运行。")
        sys.exit(0)

    keep = _parse_keep(sys.argv)

    print(f"=== Step 0: 自动备份（{len(source_files)} 个文件）===")
    stats['step0'] = step0_backup(source_files, mapping)
    stats['backup_dir'] = stats['step0']['dir']

    print(f"\n=== Step 1: 整理文件 ===")
    stats['step1'] = step1_organize(source_files, mapping)

    print(f"\n=== Step 2: 标准化格式 ===")
    stats['step2'] = step2_cleanup(stats['step1'].get('moved_paths', []))

    print(f"\n=== Step 3: 导出本批新增语料 ===")
    stats['step3_incremental'] = step3_export_incremental(
        stats['step2'].get('ok_paths', []))

    if '--no-tts' in sys.argv:
        print(f"\n=== Step 4: 自动合成（已按 --no-tts 跳过）===")
        stats['step4_tts'] = {'enabled': False}
    else:
        print(f"\n=== Step 4: 自动合成新增语料 ===")
        stats['step4_tts'] = step4_synthesize_incremental(
            stats['step3_incremental'].get('dir', ''))

    print(f"\n=== Step 5: 打包 ZIP ===")
    stats['step3'] = step3_zip(mapping=mapping)

    print(f"\n=== Step 6: 三重核验 ===")
    mapping = _build_mapping()   # 整理后重新读，确保最新
    verify_result    = step4_verify(mapping)
    stats['verify']  = verify_result

    # 排期状态回写（标记 已存在 / 本次新增 / 缺失）
    moved_names = set()
    for p in stats.get('step1', {}).get('moved_paths', []):
        e = _find_mapping_entry(os.path.splitext(os.path.basename(p))[0], mapping)
        if e:
            moved_names.add(e[4])
    status_path = os.path.join(_ensure_dir(REPORTS_DIR),
                               f"排期_状态_{RUN_TIMESTAMP}.xlsx")
    stats['schedule_status'] = write_schedule_status(mapping, moved_names, status_path)

    report_path = os.path.join(_ensure_dir(REPORTS_DIR),
                               f"处理报告_{RUN_TIMESTAMP}.txt")
    write_report(stats, report_path)

    print(f"\n=== Step 7: 最终总包 ===")
    stats['final_package'] = step5_final_package(
        stats.get('step3_incremental', {}).get('dir', ''),
        stats.get('step4_tts', {}).get('zip', ''),
        report_path,
        schedule_status=stats.get('schedule_status', ''))

    print(f"\n=== Step 8: 归档旧产物（保留最近 {keep} 份，移动到 {ARCHIVE_DIR}/）===")
    step8_prune(keep)

    if '--send' in sys.argv:
        recv = (sys.argv[sys.argv.index('--receive') + 1]
                if '--receive' in sys.argv else TRANSFER_RECEIVES)
        print(f"\n=== Step 9: 跨网段传输（{recv}）===")
        new_corpus_zip = os.path.join(OUTPUTS_DIR, f"new_corpus_{RUN_TIMESTAMP}.zip")
        tts_zip = stats.get('step4_tts', {}).get('zip', '')
        stats['step9'] = step9_transfer([new_corpus_zip, tts_zip], recv)

    print("\n流水线执行完毕。" +
          ("" if verify_result['pass'] else "  <- 请处理上方错误！"))
    sys.exit(0 if verify_result['pass'] else 1)
